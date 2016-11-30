# pyexcel_xlsx

import xlwt

from django.http import HttpResponse

def import_response(filename=None):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="%s.xlsx"' % filename
    return response


def excel_export(columns, rows):
    # response = HttpResponse(content_type='application/ms-excel')
    # response['Content-Disposition'] = 'attachment; filename="users.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Sheet1')

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    # columns = ['Username', 'First name', ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    # rows = User.objects.filter(role='user').values_list('phone', 'role')[:10]

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    return wb
    # wb.save(response)
    # return response


def export_excel(columns, rows):
    import openpyxl
    import datetime
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Sheet1', index=0)
    ws = wb.get_sheet_by_name('Sheet1')
    ws.append(columns)  # title
    # ws.column_dimensions["A"].width = 10  # 设置宽度

    for row in rows:
        print(row)
        ws.append(row)

    return wb